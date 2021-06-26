# demanda_back_end.py - Backend para leer datos históricos, ejecutar pronósticos, calcular errores, mostrar gráficos y
#                comparar entre modelos.
import copy
import datetime
import os
import shelve
import time

import matplotlib.pyplot as plt
import numpy as np
import openpyxl.styles
import pandas as pd
import pmdarima as pm
import pmdarima.arima.arima

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from statsmodels.tsa.holtwinters import ExponentialSmoothing
import statsmodels

pd.options.mode.chained_assignment = None

plt.style.use('ggplot')


def generate_testing_data():
    # generar data de prueba
    data = pd.DataFrame(np.random.randint(10, size=(100,)))
    return data


def get_excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """

    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    result = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = letters[rem]
    return ''.join(result) + str(row)


def create_excel_table_from_df(df: pd.DataFrame, sheet_: Worksheet, row_ini: 1, table_name):
    """Crea tabla de Excel en la hoja indicada a partir de un pandas DataFrame.

    Parametros:
    df: pandas DataFrame
    row_ini: fila inicial, por default 1
    sheet_: Worksheet object openpyxl
    table_name: nombre de la tabla"""

    col_last = get_excel_style(1, df.shape[1])[:-1]

    # Crear tabla de Excel
    tabla_excel = Table(displayName=table_name,
                        ref=f"A{row_ini}:{col_last}{df.shape[0] + row_ini}")  # nombre y tamaño

    # declarar estilos a la tabla
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)

    # asignar el estilo
    tabla_excel.tableStyleInfo = style

    # agregar tabla a la hoja
    sheet_.add_table(tabla_excel)


def df_to_excel(wb: Workbook, df: pd.DataFrame, sheet_: Worksheet, row_ini: 1, as_table: False, **kwargs):
    """Agregar pandas DataFrame a hoja de Excel.

    Parametros:
    df: pandas DataFrame
    sheet_: Worksheet object openpyxl
    row_ini: fila inicial, por default 1
    as_table: boolean, crear Tabla de Excel"""

    # Agregar dataframe de Python a Excel
    rows = dataframe_to_rows(df, index=False, header=True)

    # agregar filas a Excel
    for r_idx, row in enumerate(rows, row_ini):
        for c_idx, value in enumerate(row, 1):
            sheet_.cell(row=r_idx, column=c_idx, value=value)

    if as_table:
        try:
            table_name = kwargs['table_name']
            create_excel_table_from_df(df, sheet_, row_ini, table_name)
        except KeyError:
            raise ValueError('A table name must be specified if as_table is True.')
    try:
        for sheet in ['Sheet', 'Hoja', 'Hoja1']:
            wb.remove(wb[sheet])

    except KeyError:
        pass


def fill_dates_in_df(df: pd.DataFrame, date: datetime.date):
    df = df.reset_index()
    if 'Unidad_Medida' in list(df.columns):
        df = df.groupby(['Fecha', 'Codigo', 'Nombre', 'Unidad_Medida', 'Agente']).sum().reset_index()
    else:
        df = df.groupby(['Fecha', 'Codigo', 'Nombre']).sum().reset_index()

    df = df.set_index('Fecha')

    if max(df.index) != date:
        new_cols = [np.nan for x in range(df.shape[1])]
        df.loc[date] = new_cols

    # fill missing dates with 0
    df = df.asfreq('D')
    df['Demanda'] = df['Demanda'].fillna(0)
    df = df.fillna(method='ffill')

    return df


def calc_mae(data, fitted, df_index):
    df_fitted = pd.DataFrame(fitted, columns=['Ajuste'], index=df_index)
    error_df = pd.concat([data, df_fitted], axis=1)
    error_df['Error'] = abs(error_df['Ajuste'] - error_df['Demanda'])
    mae = error_df['Error'].mean()

    return mae


def change_col_sizes(sheet, cols_: list, col_len: list):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    assert len(cols_) == len(col_len)

    for i in range(len(cols_)):
        sheet.column_dimensions[letters[i]].width = col_len[i]


def timer(func):
    def wrapper(*args):
        start = time.time()

        # Call the function being decorated
        func(*args)

        #
        duration = time.time() - start
        print(f'Duration: {duration}')

    return wrapper


class FilePathShelf:
    @staticmethod
    def close_shelf(shelf: shelve):

        shelf.close()

    def __init__(self, _path):

        # path to save the shelve files
        self._path = _path

        # open shelve
        paths_shelf = shelve.open(self._path)

        # set keys list
        self._shelve_keys = ['Working',
                             'Demand',
                             'Forecast',
                             'Weighted_Forecast',
                             'BOM',
                             'Metrics_Demand',
                             'Metrics_Forecast',
                             'Temp',
                             'Demand_Agent']

        # try to get value from key, if empty initialize
        for _path in self._shelve_keys:
            try:
                paths_shelf[_path]

            except KeyError:
                paths_shelf[_path] = ''

        # close shelf
        paths_shelf.close()

    def open_shelf(self):
        paths_shelf = shelve.open(self._path)

        return paths_shelf

    def write_to_shelf(self, file_name, path_):
        """Set value (path_) to key (file_name)."""

        # open saved values
        paths_shelf = self.open_shelf()

        if file_name not in self._shelve_keys:
            raise ValueError(f'You tried to save {file_name} to the dictionary. '
                             f'The accepted values are {self._shelve_keys}.')

        # set value to key
        paths_shelf[file_name] = path_

        # save and close shelf
        self.close_shelf(paths_shelf)

    def print_shelf(self):
        """Print the shelf."""

        shelf = self.open_shelf()

        for key, value in shelf.items():
            print(key, ': ', value)

            if key is None or value is None:
                pass

        # save and close shelf
        self.close_shelf(shelf)

    def send_path(self, file_name):
        """Return path from key (file_name)."""

        paths_shelf = self.open_shelf()

        if file_name not in self._shelve_keys:
            raise ValueError(f'{file_name} is not a valid file name.')

        path = paths_shelf[file_name]

        # save and close shelf
        self.close_shelf(paths_shelf)

        return path


class ConfigShelf:

    @staticmethod
    def close_shelf(shelf: shelve):

        shelf.close()

    def __init__(self, _path):

        # path to save the shelve files
        self._path = _path

        # open shelve
        config_shelf = shelve.open(self._path)

        # set keys list
        self.config_dict = {'periods_fwd': 30,
                            'Mode': 'Demand',
                            'File_name': 'Pronóstico estadístico',
                            'File_name_agent': 'Pronóstico por agente',
                            'File_name_segmented': 'Pronóstico crítico',
                            'File_name_metrics': 'Métricas',
                            'Agg_viz': 'Diario',
                            'BOM_Explosion': False,
                            'Segmentacion': {'Supermercados': 0.4,
                                             'Panaderías': 0.3,
                                             'Restaurantes': 0.1,
                                             'Industrial': 0.05,
                                             'Abastecedor': 0.05,
                                             'Institucional': 0.02,
                                             'Particular': 0.02,
                                             'Tiquete': 0.05,
                                             'Distribuidor': 0.01}}

        # try to get value from key, if empty initialize
        for key, value in self.config_dict.items():
            try:
                config_shelf[key]
            except KeyError:
                config_shelf[key] = value

        # close shelf
        config_shelf.close()

    def open_shelf(self, writeback: bool):
        shelf = shelve.open(self._path, writeback=writeback)

        return shelf

    def write_to_shelf(self, parameter, value, **kwargs):
        """Set value (value) to key (parameter)."""

        # open saved values
        shelf = self.open_shelf(True)

        if 'model' in kwargs.keys():
            model_ = kwargs['model']

            shelf[model_]['params'][parameter][0] = value

        else:
            # set value to key
            shelf[parameter] = value

        self.config_dict = shelf

        # save and close shelf
        self.close_shelf(shelf)

    def print_shelf(self):
        """Print the shelf."""

        shelf = self.open_shelf(False)

        for key, value in shelf.items():
            print(key, ': ', value)

            if key is None or value is None:
                pass

        # save and close shelf
        self.close_shelf(shelf)

    def send_parameter(self, parameter, **kwargs):
        """Return value from key (parameter)."""

        shelf = self.open_shelf(False)

        if parameter not in shelf.keys():
            raise ValueError(f'{parameter} is not a valid parameter.')

        if 'model' in kwargs.keys():
            model_ = kwargs['model']

            value = shelf[model_]['params'][parameter][0]

        else:
            value = shelf[parameter]

        # save and close shelf
        self.close_shelf(shelf)

        return value

    def send_dict(self):
        shelf = self.open_shelf(False)

        dict_ = dict(shelf)

        self.close_shelf(shelf)

        return dict_


class Application:

    @staticmethod
    def create_total_sku_df(df: pd.DataFrame, product_codes: list, date: datetime.date):

        df_ = copy.deepcopy(df)

        df_total = pd.DataFrame()
        for unique in product_codes:
            df_sku = df_[df_['Codigo'] == unique]

            if not df_sku.empty:
                df_sku = fill_dates_in_df(df_sku, date)

            df_total = pd.concat([df_total, df_sku], axis=0)

        return df_total

    def __init__(self, path_):

        # installation path
        self.path_ = path_
        self.path_config_shelf = os.path.join(path_, 'config')
        self.path_file_paths = os.path.join(path_, 'paths')

        # initial routine
        if not self.check_if_installed():
            self.setup()

        # shelves for storing data in computer memory
        self.file_paths_shelf = FilePathShelf(self.path_file_paths)
        self.config_shelf = ConfigShelf(self.path_config_shelf)

        # master data variable
        self.raw_data = pd.DataFrame()
        self.segmented_data_sets = {}

        # available forecasting models
        self.models = {'SARIMAX': 'ARIMA'}

        # product dictionary
        self.product_dict = {}

        # products per agent dictionary
        # agents are keys and products are values, there can be multiple products (list)
        self.prods_per_agent = {}

        # parameter to define if BOM explosion should be applied or not
        self.bom_explosion = self.config_shelf.send_parameter('BOM_Explosion')

        # MODEL ATTRIBUTES

        # model chosen by the user
        self.active_model = None

        # model fitted to the data
        self.fitted_model = None

        # data used for modelling
        self.model_df = None

        # amount of periods to forecast out of sample
        self.periods_fwd = int(self.config_shelf.send_parameter('periods_fwd'))

        # feature names for the modelling data
        self.var_names = ['Demanda', 'Ajuste', 'Pronóstico']

        # possible execution modes
        self.modes = ['Demand', 'Forecast', 'Metrics', 'Demand_Agent']

        # list of available products
        self.list_product_codes = []
        self.list_product_names = []
        self.available_agents = []
        self.df_master_data = pd.DataFrame()
        self.dict_products = {}

        # dataframe with sales, with empty periods filled with 0, all products
        self.df_total_input = pd.DataFrame()

        # dataframe with real sales and values fitted by the model, all products
        self.df_total_fitted = pd.DataFrame()

        # dataframe with forecasts, all products
        self.df_total_forecasts = pd.DataFrame()

        # dataframe with all historical data and forecasts
        self.df_total_demand_fcst = pd.DataFrame()

        # dataframe with all metrics
        self.df_total_metrics = pd.DataFrame()

        # dictionary to store fitted models for forecasting
        self.dict_models_sku = {}
        self.dict_models_agent = {}

        # dictionary to store segmentation percentages for each product
        self.dict_segment_percentages = {}

        # dictionary for metric descriptions
        self.dict_metric_desc = {'AIC': ['Criterio de información de Akaike',
                                         'Utilizado para comparar modelos.'],
                                 'BIC': ['Criterio de información Bayesiano',
                                         'Utilizado para comparar modelos.'],
                                 'Bias': ['Sesgo',
                                          'Es el promedio del error un valor positivo indica\n'
                                          'una sobreestimación de la demanda y viceversa.'],
                                 'Bias_PERC': ['Sesgo Porcentual',
                                               'Indica el sesgo como proporción de la\n'
                                               'demanda promedio.'],
                                 'MAE': ['Error absoluto medio (MAE)',
                                         'Es el promedio del error absoluto, indica el valor\n'
                                         'promedio del error en la unidad de medida de los\n'
                                         'datos de entrada.'],
                                 'MAE_PERC': ['MAE Porcentual',
                                              'Indica el error promedio como proporción de la\n'
                                              'demanda promedio.'],
                                 'MSE': ['Error cuadrático medio (MSE)',
                                         'Es el promedio del error cuadrático, indica el\n'
                                         'valor promedio del error elevado al cuadrado.'],
                                 'RMSE': ['Raíz del error cuadrático medio (RMSE)',
                                          'Indica la raíz de MSE en la unidad de medida '
                                          'de los datos de entrada.'],
                                 'RMSE_PERC': ['RMSE Porcentual',
                                               'Indica el RMSE como proporción de la demanda\n'
                                               ' promedio.']}

    def setup(self):
        if not os.path.exists(self.path_):
            print('Instalando el programa.')
            os.makedirs(self.path_)

    def check_if_installed(self):

        if os.path.exists(self.path_):
            return True

        else:
            return False

    def set_path(self, filename, path):

        self.file_paths_shelf.write_to_shelf(filename, path)

    def get_path(self, filename):

        return self.file_paths_shelf.send_path(filename)

    def set_parameter(self, parameter, value):

        self.config_shelf.write_to_shelf(parameter, value)

    def get_parameter(self, parameter):

        return self.config_shelf.send_parameter(parameter)

    def clear_data(self):
        """Clear information from the back end."""

        self.dict_fitted_dfs = {}
        self.dict_errors = {}

    def read_data(self, process_: str):
        """Returns pandas dataframe with time series data."""

        # Get Demand path from parameters shelf
        path = self.file_paths_shelf.send_path(process_)

        mapping = {'Demand': 'demanda',
                   'Forecast': 'pronóstico',
                   'Metrics_Demand': 'demanda',
                   'Metrics_Forecast': 'pronóstico',
                   'Demand_Agent': 'demanda'}

        # raise value error if the key is empty
        if path == '':
            err = f"El directorio hacia el archivo de {mapping[process_]} no esta definido."
            raise KeyError(err)

        # if file ends with CSV, read as CSV
        if path.endswith('.csv'):
            df = pd.read_csv(path, sep=",", decimal=",", header=0)
            return df

        # if file ends with xlsx, read as Excel
        elif path.endswith('.xlsx'):
            df = pd.read_excel(path)
            return df

    def read_data_from_path(self, path_name):
        """Returns pandas dataframe with time series data."""

        # Get Demand path from parameters shelf
        path = self.file_paths_shelf.send_path(path_name)

        # raise value error if the key is empty
        if path_name == '':
            err = f"El directorio  {path_name} no esta definido."
            raise KeyError(err)

        # if file ends with CSV, read as CSV
        if path.endswith('.csv'):
            df = pd.read_csv(path, sep=",", decimal=",", header=0)
            return df

        # if file ends with xlsx, read as Excel
        elif path.endswith('.xlsx'):
            df = pd.read_excel(path)
            return df

    def clean_data(self, process_: str):
        """Cleans the time series data.
        First column is assumed to have datetime like values.
        Second column is assumed to be SKU.
        Third column is assumed to be the name of the SKU.
        Last column is assumed to be the demand values, numerical.
        Columns in between the third and the last are treated as extra aggregation parameters for the forecast."""

        # read the data
        df = self.read_data(process_)

        # Metrics forecast could be two different files, test first
        if process_ == 'Metrics_Forecast':
            if df.shape[1] == 4:
                process_ = 'Metrics_Forecast_Stats'
            else:
                process_ = 'Metrics_Forecast_Colab'

        # Dictionary for each process
        # The item is another dictionary with the column mapping for each process
        col_mapping = {'Demand': ['Fecha',
                                  'Codigo',
                                  'Nombre',
                                  'Demanda'],
                       'Forecast': ['Fecha',
                                    'Codigo',
                                    'Nombre',
                                    'Pronóstico'],
                       'Metrics_Demand': ['Fecha',
                                          'Codigo',
                                          'Nombre',
                                          'Demanda'],
                       'Metrics_Forecast_Stats': ['Fecha',
                                                  'Codigo',
                                                  'Nombre',
                                                  'Pronóstico'],
                       'Metrics_Forecast_Colab': ['Fecha',
                                                  'Codigo',
                                                  'Nombre',
                                                  'Grupo',
                                                  'Pronóstico'],
                       'Demand_Agent': ['Fecha',
                                        'Codigo',
                                        'Nombre',
                                        'Unidad_Medida',
                                        'Agente',
                                        'Demanda']}
        file_name = col_mapping[process_][-1]
        cols = col_mapping[process_]

        if df.shape[1] != len(cols):
            raise ValueError(f'El archivo de {file_name} indicado tiene una estructura incorrecta.\n'
                             f'Se requieren {len(cols)} columnas {"-".join(cols)}.\n'
                             f'El archivo cargado tiene {df.shape[1]} columnas.')

        df.columns = col_mapping[process_]

        # convert first column to datetime or raise ValueError
        try:
            df['Fecha'] = pd.to_datetime(df['Fecha'])
        except ValueError:
            raise ValueError('Error: en la primera columna de los datos de entrada hay filas que no contienen fechas.')

        # convert last column to numerical or raise ValueError
        try:
            df.iloc[:, -1] = pd.to_numeric(df.iloc[:, -1])
        except ValueError:
            raise ValueError('Error: en la ultima columna de los datos de entrada hay filas que no contienen'
                             ' datos numericos.')

        # extract date from datetime values
        df['Fecha'] = df['Fecha'].dt.date

        # group demand by date and categorical features (sum)
        df = df.groupby(list(df.columns[:-1])).sum().reset_index()

        # Ordenar por fecha
        df = df.sort_values(['Fecha'])

        # set date as index
        df.set_index(['Fecha'], inplace=True)
        df.index = pd.DatetimeIndex(df.index)

        # save df as a class attribute
        self.raw_data = df

        return df

    def apply_bom(self, df: pd.DataFrame, process: str):
        """Convert the final product demand to its base components using a BOM (Bill of materials)."""

        # BOM path
        path_bom = self.file_paths_shelf.send_path('BOM')

        # raise value error if the key is empty
        if path_bom == '':
            err = "El directorio hacia el archivo de recetas no esta definido."
            raise KeyError(err)

        # group original demand data by selected fields
        df = df.groupby(['Fecha', 'Codigo', 'Nombre']).sum().reset_index()
        df.columns = ['Fecha', 'Cod_Prod', 'Nombre_Prod', 'Cantidad']

        # read BOM file
        bom = pd.read_excel(path_bom)

        if bom.shape[1] != 6:
            raise ValueError('El archivo de recetas debe tener 6 columnas:\nCódigo Receta, Nombre Receta,'
                             'Cantidad Receta', 'Código Artículo', 'Nombre Artículo', 'Cantidad Artículo')

        # select columns and change column names
        bom.columns = ['Cod_Prod',
                       'Nombre_Prod',
                       'Cant_Prod',
                       'Cod_Comp',
                       'Nombre_Comp',
                       'Cant_Comp']

        # create table with intermediate products, that are not final components
        intermediate = pd.DataFrame(bom['Cod_Comp'].drop_duplicates()).dropna()
        intermediate.columns = ['Cod_Inter']

        # join intermediate products with BOM to get conversion factors between
        # intermediate products and final components
        intermediate = intermediate.merge(bom, left_on='Cod_Inter', right_on='Cod_Prod', how='left')
        intermediate.rename(columns={'Nombre_Prod': 'Nombre_Inter'}, inplace=True)
        intermediate.drop(columns=['Cod_Prod'], inplace=True)
        intermediate.dropna(subset=['Cod_Comp'], inplace=True)

        # apply the BOM explosion to the original demand data
        df_bom = df.merge(bom.drop(columns=['Nombre_Prod']), on='Cod_Prod', how='left')
        df_bom['Cant_Req'] = df_bom['Cantidad'] * df_bom['Cant_Comp']

        # group the new data by the component demand
        df_bom = df_bom.groupby(['Cod_Comp', 'Nombre_Comp', 'Fecha'])['Cant_Req'].sum().reset_index()
        df_bom = df_bom[~df_bom['Nombre_Comp'].str.contains('RECORTES')]
        df_bom.columns = ['Cod_Prod', 'Nombre_Prod', 'Fecha', 'Cant_Req']

        # apply the BOM explosion to the dataset, to the get the demand for the final components
        df_bom = df_bom.merge(intermediate, left_on='Cod_Prod', right_on='Cod_Inter', how='left')
        df_bom.loc[df_bom['Cod_Inter'].notna(),
                   'Cant_Req_Final'] = df_bom['Cant_Req'] * df_bom['Cant_Comp']
        df_bom.loc[df_bom['Cod_Inter'].notna(), 'Codigo'] = df_bom['Cod_Comp']
        df_bom.loc[df_bom['Cod_Inter'].notna(), 'Nombre'] = df_bom['Nombre_Comp']

        df_bom.loc[df_bom['Cant_Req_Final'].isna(), 'Codigo'] = df_bom[
            'Cod_Prod']
        df_bom.loc[df_bom['Cant_Req_Final'].isna(), 'Nombre'] = df_bom[
            'Nombre_Prod']
        df_bom.loc[df_bom['Cant_Req_Final'].isna(), 'Cant_Req_Final'] = df_bom[
            'Cant_Req']

        # keep selected columns and reorder
        df_bom = df_bom[['Codigo', 'Nombre', 'Fecha', 'Cant_Req_Final']]

        if process == 'Forecast':
            df_bom.columns = ['Codigo', 'Nombre', 'Fecha', 'Pronóstico']
        else:
            df_bom.columns = ['Codigo', 'Nombre', 'Fecha', 'Demanda']

        # extract date from datetime values
        df_bom['Fecha'] = df_bom['Fecha'].dt.date

        # group demand by date and categorical features (sum)
        df_bom = df_bom.groupby(['Fecha', 'Codigo', 'Nombre']).sum().reset_index()

        # set date as index
        df_bom.set_index(['Fecha'], inplace=True)
        df_bom.index = pd.DatetimeIndex(df_bom.index)

        self.raw_data = df_bom

        # return dataset
        return df_bom

    def create_metrics_df(self, df_demand: pd.DataFrame, df_forecast: pd.DataFrame):
        """
        Receives a data frame with the real demand of a period and the forecast of the same period.
        Returns a data frame with demand and forecast as columns to be able to compare both.
        """

        # Group the forecast by date, product code and product name, sum values.
        df_forecast = df_forecast.groupby(['Fecha', 'Codigo', 'Nombre'])['Pronóstico'].sum().reset_index()

        # Create table of codes and names in case names dont match
        df_demand_names = df_demand[['Codigo', 'Nombre']]
        df_fcst_names = df_forecast[['Codigo', 'Nombre']]
        df_names = pd.concat([df_demand_names, df_fcst_names])
        df_names.drop_duplicates(subset=['Codigo'], keep='first', inplace=True)

        # Create joint table, using date and product code as keys, outer join.
        df_demand.drop(columns=['Nombre'], inplace=True)
        df_forecast.drop(columns=['Nombre'], inplace=True)
        df = df_demand.merge(df_forecast, on=['Fecha', 'Codigo'], how='outer')

        # Add the product name to the table using the product code as key, left join on the demand-forecast table.
        df = df.merge(df_names, on='Codigo', how='left')

        # Change column order
        df = df[['Fecha',
                 'Codigo',
                 'Nombre',
                 'Demanda',
                 'Pronóstico']]

        # Fill NaN values
        for col in ['Demanda', 'Pronóstico']:
            df[col] = df[col].fillna(0)

        # Calculate the forecast error
        df['Error'] = df['Pronóstico'] - df['Demanda']

        # Set date as index
        df.set_index(['Fecha'], inplace=True)
        df.index = pd.DatetimeIndex(df.index)

        # Calculate bias and MAE for each product and add them to dataframe for exporting.
        df_export = copy.deepcopy(df)

        df_export_demand_fcst = df_export.groupby(['Codigo', 'Nombre'])[['Demanda', 'Pronóstico']].sum().reset_index()

        df_export['Abs_Error'] = df_export['Error'].abs()
        df_export_bias = df_export.groupby(['Codigo', 'Nombre'])['Error'].mean().reset_index()
        df_export_mae = df_export.groupby(['Codigo', 'Nombre'])['Abs_Error'].mean().reset_index()
        df_export = df_export_bias.merge(df_export_mae, on=['Codigo', 'Nombre'], how='left')
        df_export = df_export.merge(df_export_demand_fcst, on=['Codigo', 'Nombre'], how='left')
        df_export['MAE (%)'] = df_export['Abs_Error'] / df_export['Demanda']
        df_export['Sesgo (%)'] = df_export['Error'] / df_export['Demanda']

        df_export.columns = ['Codigo', 'Nombre', 'Sesgo', 'MAE', 'Demanda', 'Pronóstico', 'MAE (%)', 'Sesgo (%)']
        df_export = df_export[['Codigo', 'Nombre', 'Demanda', 'Pronóstico', 'MAE', 'Sesgo', 'MAE (%)', 'Sesgo (%)']]

        self.df_error_export = df_export

        # Save data frame to class attribute.
        self.raw_data = df

        return df

    @staticmethod
    def create_master_data_df(df, columns):
        """Create a dataframe with unique values for each of the columns passed."""

        temp_df = pd.DataFrame(df['Codigo'].unique(), columns=['Codigo'])
        temp_df = temp_df.merge(df[columns], on='Codigo', how='left').drop_duplicates(subset=['Codigo'])

        return temp_df

    def create_input_df(self, process_: str, apply_bom: bool):
        """Separate the raw data into N datasets, where N is the number of unique products in the raw data."""

        # Clean data upon function call, must read and clean two files
        if process_ == 'Metrics':
            df_metrics_demand = self.clean_data('Metrics_Demand')
            df_metrics_demand = self.apply_bom(df_metrics_demand, process_)
            df_metrics_fcst = self.clean_data('Metrics_Forecast')
            df_input = self.create_metrics_df(df_metrics_demand, df_metrics_fcst)

        else:
            df_input = self.clean_data(process_)

            # If bom_explosion is True, apply the BOM Explosion to the raw data
            if self.config_shelf.send_parameter('BOM_Explosion') and process_ == 'Demand':
                df_input = self.apply_bom(df_input, process_)

        if apply_bom:
            df_input = self.apply_bom(df_input, process_)

        # Create dataframe with master data for products, agents
        if process_ == 'Demand_Agent':
            cols = ['Codigo', 'Nombre', 'Unidad_Medida']
        else:
            cols = ['Codigo', 'Nombre']
        self.df_master_data = self.create_master_data_df(df_input, cols)

        # Create two lists of unique product codes and names, combine both in a list
        self.list_product_codes = [code for code in self.df_master_data.loc[:, 'Codigo'].unique()]
        self.list_product_names = [code for code in self.df_master_data.loc[:, 'Nombre'].unique()]

        self.dict_products = dict(zip(self.list_product_codes, self.list_product_names))

        # If running the Demand process, reindex to fill missing dates with Demand 0.
        if process_ == 'Demand':
            df_input = df_input.drop(columns=['Nombre']).reset_index(). \
                merge(self.df_master_data, on='Codigo', how='left').set_index(['Fecha'])

            df_input = self.create_total_sku_df(df_input, self.list_product_codes, datetime.date.today())

        if process_ == 'Demand_Agent':
            # If running the Demand_Agent process, create a list of all the unique agents.

            self.available_agents = [agent for agent in df_input.loc[:, 'Agente'].unique()]
            df_input.drop(columns=['Unidad_Medida'], inplace=True)

            temp_df = pd.DataFrame()

            for agent in self.available_agents:
                df_agent = df_input[df_input['Agente'] == agent]

                df_agent = df_agent.drop(columns=['Nombre']).reset_index(). \
                    merge(self.df_master_data, on='Codigo', how='left').set_index(['Fecha'])
                df_agent = self.create_total_sku_df(df_agent, self.list_product_codes, datetime.date.today())

                self.prods_per_agent[agent] = list(df_agent['Nombre'].unique())

                temp_df = pd.concat([temp_df, df_agent], axis=0)

            df_input = copy.deepcopy(temp_df)

        # assign the dictionary to class attribute
        self.df_total_input = df_input

    @staticmethod
    def fit_model(demand_col, df_index):

        # get the best ARIMA model for each df
        arima_model = pm.auto_arima(demand_col,
                                    out_of_sample_size=20,
                                    stepwise=True)

        # fit the best model to the dataset
        arima_model = arima_model.fit(demand_col)
        arima_mae = calc_mae(demand_col, arima_model.arima_res_.fittedvalues, df_index)

        base_mae = arima_mae

        # create a dataset with the fitted values
        df_fitted = pd.DataFrame(arima_model.arima_res_.fittedvalues, columns=['Ajuste'], index=df_index)
        model = arima_model

        # get a Winter's Exponential smoothing model and calculate it's MAE
        holt_model = ExponentialSmoothing(demand_col, freq='D')
        holt_model = holt_model.fit()
        holt_mae = calc_mae(demand_col, holt_model.fittedvalues, df_index)
        if holt_mae < base_mae:
            model = holt_model
            df_fitted = pd.DataFrame(holt_model.fittedvalues, columns=['Ajuste'], index=df_index)

        return model, df_fitted

    def fit_model_sku_list(self, queue_, df_total_input, sku_list, sku_amount: int, start: int):
        """Get an optimized model for each of the separate product data sets."""

        # check if data is loaded
        if df_total_input.empty:
            raise ValueError('No hay datos cargados para crear un modelo.')

        # check amount of data sets to use as a way of measuring progress bar
        # num_skus = len(sku_list)
        print(f'Cantidad de SKUs: {sku_amount}')

        # iterate over data sets for training and predictions
        df_fitted_skus = pd.DataFrame()
        dict_fitted_models_sku = {}
        for idx, sku in enumerate(sku_list, start):
            queue_.put([f'Entrenando modelo para {self.dict_products[sku]}.\n',
                        0])

            start = time.time()

            df_sku = df_total_input[df_total_input['Codigo'] == sku]

            # create a dataset with the real data
            df_sku_input = pd.DataFrame(df_sku.loc[:, 'Demanda'], columns=['Demanda'])

            model, df_sku_fitted = self.fit_model(df_sku.loc[:, 'Demanda'], df_sku_input.index)

            # join the real data with the fitted values on the rows axis
            df_sku_fitted = pd.concat([df_sku, df_sku_fitted], axis=1)

            # Add model to dictionary of models
            dict_fitted_models_sku[sku] = model

            df_fitted_skus = pd.concat([df_fitted_skus, df_sku_fitted], axis=0)

            duration = round(time.time() - start, 2)
            queue_.put([f'Modelo para {self.dict_products[sku]} listo.', idx / sku_amount])

            queue_.put([f'Progreso : {idx}/{sku_amount}.\n', 0])

            queue_.put([f'Duración: {duration} segundos.', 0])

            queue_.put(['', 0])

            print(f'Progreso: {idx}/{sku_amount}. Porcentaje equivale a: {idx / sku_amount}')

            print(f'Duración para {self.dict_products[sku]} {time.time() - start} s.')

        # queue_.put(['', percentage])

        return df_fitted_skus, dict_fitted_models_sku

    def forecast(self, df, fitted_model):
        """Predict N periods forward using self.periods_fwd as N."""

        periods_fwd = self.config_shelf.send_parameter('periods_fwd')

        # create index from the max date in the original dataset to periods_fwd days forward
        pred_index = pd.date_range(start=df.index.max() + datetime.timedelta(days=1),
                                   end=df.index.max() + datetime.timedelta(days=periods_fwd))

        if type(fitted_model) == pmdarima.arima.arima.ARIMA:

            # get OOB forecast and confidence intervals
            predictions, confidence = fitted_model.predict(n_periods=periods_fwd,
                                                           return_conf_int=True)

            # add the confidence interval (both bounds) to a DataFrame using the prediction index
            confidence = pd.DataFrame(confidence, index=pred_index, columns=['Lower', 'Upper'])

        else:
            # If using Holt Winters use simulations to get the confidence intervals
            predictions = fitted_model.forecast(periods_fwd)
            confidence = fitted_model.simulate(periods_fwd, repetitions=100)
            confidence['Lower'] = confidence.min(axis=1)
            confidence['Upper'] = confidence.max(axis=1)
            confidence = confidence[['Lower', 'Upper']]

        # add the predictions to a DataFrame using the prediction index
        predictions = pd.DataFrame(predictions, index=pred_index, columns=[self.var_names[0]])

        # allow only non-negative predictions
        predictions.loc[predictions[self.var_names[0]] < 0, self.var_names[0]] = 0

        # concatenate the base predictions and the confidence interval to get a three-column DataFrame
        full_preds = pd.concat([predictions, confidence], axis=1)

        return full_preds

    def forecast_sku_list(self, df_fitted_skus, sku_list, dict_fitted_models_sku):

        # df_total = copy.deepcopy(self.df_total_fitted)
        df_forecast_skus = pd.DataFrame()
        df_demand_forecast_skus = pd.DataFrame()

        # check if data is loaded
        if df_fitted_skus.empty:
            raise ValueError('No hay datos cargados para crear un modelo.')

        # iterate over data sets for training and predictions
        for sku in sku_list:
            df_sku = df_fitted_skus[df_fitted_skus['Codigo'] == sku]

            # create a dataset with the real data
            df_sku_input = pd.DataFrame(df_sku.loc[:, 'Demanda'], columns=['Demanda'])

            # Get trained model with the sku key
            model = dict_fitted_models_sku[sku]

            # call a function to get an out of sample prediction, result is a dataset with predictions
            df_forecast_sku = self.forecast(df_sku_input, model)
            df_forecast_sku.columns = ['Pronóstico', 'Min', 'Max']

            df_sku_forecast = copy.deepcopy(df_forecast_sku)
            df_sku_forecast['Codigo'] = sku
            df_sku_forecast['Nombre'] = self.dict_products[sku]
            df_sku_forecast = df_sku_forecast[['Codigo', 'Nombre', 'Pronóstico', 'Min', 'Max']]

            df_forecast_skus = pd.concat([df_forecast_skus, df_sku_forecast], axis=0)
            df_forecast_skus = df_forecast_skus.ffill()

            # concat the predictions to the (data, fitted) dataset to get all values in one dataset
            df_sku_demand_fcst = pd.concat([df_sku, df_sku_forecast], axis=0)
            df_demand_forecast_skus = pd.concat([df_demand_forecast_skus, df_sku_demand_fcst], axis=0)

            try:
                df_demand_forecast_skus['Agente'] = df_demand_forecast_skus['Agente'].ffill()
            except KeyError:
                pass

        return df_forecast_skus, df_demand_forecast_skus

    def create_metrics_df_sku(self, sku, df_fitted_sku):
        """
        Creates a DataFrame with calculated metrics as columns and the SKU as the unique row.

        Args:
            sku: The SKU (product code).
            df_fitted_sku: Dataframe with demand and forecast as columns for the SKU.

        Returns:
            A DataFrame with each metric as a column and the SKU as a row.
        """

        # Create a copy of the DF to modify it
        df = copy.deepcopy(df_fitted_sku)

        # error = forecast - demand
        df.loc[:, 'Error'] = df['Ajuste'] - df['Demanda']

        # absolute error = abs(error)
        df.loc[:, 'Abs_Error'] = df['Error'].abs()

        # squared error
        df.loc[:, 'Squared_Error'] = df['Error'] ** 2

        # calculate the bias, mean of error
        bias = df['Error'].mean()

        # calculate the bias as a percentage of total demand
        bias_perc = (bias / df[self.var_names[0]].mean()) * 100

        # calculate the mean absolute error, mean of absolute error
        mae = df['Abs_Error'].mean()

        # calculate the mean percentage absolute error
        mae_perc = (mae / df[self.var_names[0]].mean()) * 100

        # calculate the mean squared error
        mse = df['Squared_Error'].mean()

        # calculate the rmse
        rmse_ = mse ** (1 / 2)

        # calculate the rmse percentage
        rmse_perc = (rmse_ / df[self.var_names[0]].mean()) * 100

        # create the dataframe
        dict_metrics = {'Codigo': [sku],
                        'Bias': [bias],
                        'Bias_PERC': [bias_perc],
                        'MAE': [mae],
                        'MAE_PERC': [mae_perc],
                        'MSE': [mse],
                        'RMSE': [rmse_],
                        'RMSE_PERC': [rmse_perc]}
        df_metrics = pd.DataFrame(dict_metrics)

        return df_metrics

    def calculate_metrics_sku_list(self, df_fitted_skus: pd.DataFrame, sku_list: list):
        """
        Creates a DataFrame with metrics for each SKU in sku_list and concatenates to a master DataFrame.

        Args:
            df_fitted_skus: Dataframe with demand and forecast as columns for all SKUs.
            sku_list = list of all SKUs

        Returns:
            Master DataFrame with all the metrics for each SKU.
        """

        # empty dataframe, to concat every created dataframe
        df_metrics_skus = pd.DataFrame()

        # for each sku in the list
        for sku in sku_list:
            # get the fitted values dataframe by slicing
            df_fitted_sku = df_fitted_skus[df_fitted_skus['Codigo'] == sku]

            # create the metrics dataframe
            df_metrics_sku = self.create_metrics_df_sku(sku, df_fitted_sku)

            # concatenate the sku metrics df to the the total df
            df_metrics_skus = pd.concat([df_metrics_skus, df_metrics_sku], axis=0)

        return df_metrics_skus

    def fit_forecast_evaluate_pipeline(self, process: str, queue_):

        df_total_input = copy.deepcopy(self.df_total_input)

        queue_.put([f'Comenzando proceso de entrenamiento.\n', 0])

        if process == 'Demand':
            sku_list = self.list_product_codes

            df_fitted_skus, dict_fitted_models_sku = self.fit_model_sku_list(queue_,
                                                                             df_total_input,
                                                                             sku_list,
                                                                             len(sku_list),
                                                                             1)

            df_forecast_skus, df_demand_forecast_skus = self.forecast_sku_list(df_fitted_skus,
                                                                               sku_list,
                                                                               dict_fitted_models_sku)

            df_metrics_skus = self.calculate_metrics_sku_list(df_fitted_skus,
                                                              sku_list)

            self.df_total_fitted = df_fitted_skus
            self.df_total_forecasts = df_forecast_skus
            self.df_total_demand_fcst = df_demand_forecast_skus
            self.df_total_metrics = df_metrics_skus
            self.dict_models_sku = dict_fitted_models_sku

        if process == 'Demand_Agent':

            df_fitted_agents = pd.DataFrame()
            df_forecast_agents = pd.DataFrame()
            df_demand_forecast_agents = pd.DataFrame()
            df_metrics_agents = pd.DataFrame()
            dict_models_agent = {}
            num_agents = len(self.available_agents)

            # get total sku amount
            sku_amount = 0
            for agent in self.available_agents:
                df_agent = df_total_input[df_total_input['Agente'] == agent]
                sku_list_agent = list(df_agent['Codigo'].unique())
                sku_amount += len(sku_list_agent)

            acum = 1
            for idx, agent in enumerate(self.available_agents):
                queue_.put([f'Entrenando modelos para {agent}.\n', 0])
                print(f'Entrenando modelos para {agent}.\n')

                df_agent = df_total_input[df_total_input['Agente'] == agent]

                sku_list_agent = list(df_agent['Codigo'].unique())

                df_fitted_skus, dict_fitted_models_sku = self.fit_model_sku_list(queue_,
                                                                                 df_agent,
                                                                                 sku_list_agent,
                                                                                 sku_amount,
                                                                                 acum)
                acum += len(sku_list_agent)

                df_fitted_agents = pd.concat([df_fitted_agents, df_fitted_skus], axis=0)
                dict_models_agent[agent] = dict_fitted_models_sku

                df_forecast_skus, df_demand_forecast_skus = self.forecast_sku_list(df_fitted_skus,
                                                                                   sku_list_agent,
                                                                                   dict_fitted_models_sku)
                df_forecast_skus['Agente'] = agent
                df_demand_forecast_skus['Agente'] = df_demand_forecast_skus['Agente'].ffill()

                df_forecast_agents = pd.concat([df_forecast_agents,
                                                df_forecast_skus],
                                               axis=0)
                df_demand_forecast_agents = pd.concat([df_demand_forecast_agents,
                                                       df_demand_forecast_skus],
                                                      axis=0)

                df_metrics_skus = self.calculate_metrics_sku_list(df_fitted_skus,
                                                                  sku_list_agent)
                df_metrics_skus['Agente'] = agent
                df_metrics_agents = pd.concat([df_metrics_agents,
                                               df_metrics_skus],
                                              axis=0)

                queue_.put([f'Modelo para {agent} listo.\n', idx / num_agents])
                queue_.put(['', 0])

            queue_.put([f'Proceso de entrenamiento terminado.\n', 0])
            self.df_total_fitted = df_fitted_agents
            self.df_total_forecasts = df_forecast_agents
            self.df_total_demand_fcst = df_demand_forecast_agents
            self.df_total_metrics = df_metrics_agents
            self.dict_models_agent = dict_models_agent

    def refresh_predictions(self, process):

        if self.df_total_fitted.empty:
            raise ValueError('Debe entrenar los modelos primero.')

        if process in ['Demand', 'Model']:

            sku_list = self.list_product_codes
            df_fitted_skus = self.df_total_fitted
            dict_models_sku = self.dict_models_sku

            df_forecast_skus, df_demand_forecast_skus = self.forecast_sku_list(df_fitted_skus,
                                                                               sku_list,
                                                                               dict_models_sku)

            self.df_total_forecasts = df_forecast_skus
            self.df_total_demand_fcst = df_demand_forecast_skus

        else:

            df_forecast_agents = pd.DataFrame()
            df_fitted_agents = copy.deepcopy(self.df_total_fitted)
            df_demand_forecast_agents = pd.DataFrame()
            for agent in self.available_agents:
                sku_list_agent = list(self.dict_models_agent[agent].keys())
                df_fitted_skus = df_fitted_agents[df_fitted_agents['Agente'] == agent]
                dict_fitted_models_sku = self.dict_models_agent[agent]

                df_forecast_skus, df_demand_forecast_skus = self.forecast_sku_list(df_fitted_skus,
                                                                                   sku_list_agent,
                                                                                   dict_fitted_models_sku)

                df_forecast_skus['Agente'] = agent

                df_forecast_agents = pd.concat([df_forecast_agents,
                                                df_forecast_skus],
                                               axis=0)
                df_demand_forecast_agents = pd.concat([df_demand_forecast_agents,
                                                       df_demand_forecast_skus],
                                                      axis=0)

            self.df_total_forecasts = df_forecast_agents
            self.df_total_demand_fcst = df_demand_forecast_agents

    def disaggregate_forecast_workflow(self, method='Weighted_Forecast'):

        df_to_disaggregate = copy.deepcopy(self.df_total_input)

        if method == 'Weighted_Forecast':

            df_disaggregation = self.read_data_from_path(method)
            df_disaggregated = self.disaggregate_forecast_with_df(df_to_disaggregate,
                                                                  df_disaggregation)

        else:
            disaggregation_dict = self.get_parameter('Segmentacion')
            df_disaggregated = self.disaggregate_forecast_with_dict(df_to_disaggregate,
                                                                    'Pronóstico',
                                                                    'Grupo',
                                                                    disaggregation_dict)

        return df_disaggregated

    def disaggregate_forecast_with_df(self, df_to_disaggregate: pd.DataFrame, df_disaggregation: pd.DataFrame):

        # Create copies to edit
        df_to_disaggregate_local = copy.deepcopy(df_to_disaggregate)
        df_disaggregation_local = copy.deepcopy(df_disaggregation)

        # Change column names
        df_disaggregation_local.columns = ['Fecha',
                                           'Ruta',
                                           'Codigo',
                                           'Nombre',
                                           'Pronostico_disagg',
                                           'Unidad_Medida']

        # Drop unnecessary columns
        df_disaggregation_local = df_disaggregation_local[['Fecha', 'Ruta', 'Codigo', 'Pronostico_disagg']]

        # Reset index and drop unnecessary columns
        df_to_disaggregate_local = df_to_disaggregate_local.reset_index()
        df_to_disaggregate_local.columns = ['Fecha', 'Codigo', 'Nombre', 'Pronostico_agg']

        # Create disaggregated dataframe
        df_disaggregated = df_to_disaggregate_local.merge(df_disaggregation_local, on=['Fecha', 'Codigo'], how='outer')

        # Multiply aggregate forecast with disaggregate forecast to disaggregate the aggregate forecast
        df_disaggregated['Pronóstico'] = df_disaggregated['Pronostico_agg'] * df_disaggregated['Pronostico_disagg']

        # Drop unnecessary columns
        df_disaggregated = df_disaggregated[['Fecha',
                                             'Ruta',
                                             'Codigo',
                                             'Nombre',
                                             'Pronóstico']]

        return df_disaggregated

    def disaggregate_forecast_with_dict(self, df: pd.DataFrame, col_to_disaggregate: str, disaggregation_column: str,
                                        disaggregation_dict: dict):

        df_disaggregated = pd.DataFrame()

        # For each sales group, get a new dataframe and concatenate each to an empty one.
        for key, value in disaggregation_dict.items():
            df[disaggregation_column] = key
            df[col_to_disaggregate] = df[col_to_disaggregate] * float(value)
            df_disaggregated = pd.concat([df_disaggregated, df], axis=0)

        return df_disaggregated

    def weight_forecast(self, df):

        df_grouped = df.groupby(['Fecha', 'Agente'])['Pronóstico'].sum().reset_index()
        df_grouped.columns = ['Fecha', 'Agente', 'Total']
        df_proportions = df.merge(df_grouped, on=['Fecha', 'Agente'], how='left')
        df_proportions['Pronóstico %'] = df_proportions['Pronóstico'] / df_proportions['Total']
        df_proportions = df_proportions.fillna(0)

        df_proportions.drop(columns=['Total', 'Pronóstico'], inplace=True)

        return df_proportions

    def export_data(self, path, file_name, extension, process, **kwargs):
        """
        Callback for the Export button from the GUI.
        Exports the relevant data depending on the process parameter."""

        # Define the filename
        file_name = file_name + extension

        # If the process is demand or demand agent
        # Check if the forecast has been executed, if not raise errors.
        if process in ['Demand', 'Demand_Agent', 'Model', 'Model_Agent']:
            if self.df_total_forecasts.empty:
                raise ValueError('The model has to be trained first.')
            else:
                df = self.df_total_forecasts

        elif kwargs.keys().__contains__('df'):
            df = kwargs['df']

        # If the process is Forecast or Metrics, use the input as the df to export.
        else:
            df = self.df_total_input

        # Reset the index to get the date as a column
        df = df.reset_index()

        # Rename the date column
        try:
            df = df.rename(columns={'index': 'Fecha'})
        except KeyError:
            pass

        # Get the date from the datetime values
        df['Fecha'] = df['Fecha'].dt.date

        # --- COLUMN MAPPING AND SIZES ---

        # If process is Demand, 6 columns.
        if process in ['Demand', 'Model']:
            col_order = ['Fecha',
                         'Codigo',
                         'Nombre',
                         'Pronóstico',
                         'Min',
                         'Max']
            col_sizes = [12, 12, 40, 12, 12, 12]

        # If process is Demand Agent, 11 columns.
        elif process == 'Demand_Agent':

            if kwargs['weighted_forecast']:
                df = self.weight_forecast(df)

            # Add the master data to the forecast DF
            df = df.merge(self.df_master_data[['Codigo', 'Unidad_Medida']], on='Codigo', how='left')

            # Add extra columns
            # df['Fecha creacion'] = datetime.date.today().strftime('%d-%m-%Y')
            # df['Codigo cliente'] = 'ESTIMADO'
            # df['Nombre cliente'] = 'Estimado por agente'

            # Rename columns
            df = df.rename(columns={'Codigo': 'Codigo producto',
                                    'Nombre': 'Nombre producto'})

            # Column order
            '''col_order = ['Fecha creacion',
                         'Fecha',
                         'Agente',
                         'Codigo producto',
                         'Nombre producto',
                         'Codigo cliente',
                         'Nombre cliente',
                         'Pronóstico %',
                         'Unidad_Medida',
                         'Min',
                         'Max']'''

            col_order = ['Fecha',
                         'Agente',
                         'Codigo producto',
                         'Nombre producto',
                         'Pronóstico %',
                         'Unidad_Medida',
                         'Min',
                         'Max']
            # Column sizes
            # col_sizes = [12, 12, 12, 12, 40, 12, 40, 12, 12, 12, 12]
            col_sizes = [12, 12, 12, 40, 12, 12, 12, 12]

        # If process is Forecast, 5 columns.
        elif process == 'Forecast':
            col_order = ['Fecha',
                         'Codigo',
                         'Nombre',
                         'Grupo',
                         'Pronóstico']
            col_sizes = [12, 12, 40, 12, 12]

        # If process is Metrics, 6 columns.
        else:
            col_order = ['Fecha',
                         'Codigo',
                         'Nombre',
                         'Demanda',
                         'Pronóstico',
                         'Error']
            col_sizes = [12, 12, 40, 12, 12, 12]

        # Change column order
        df = df[col_order]

        if extension == '.xlsx':

            wb = Workbook()
            sheet = wb.create_sheet('Pronóstico')

            df_to_excel(wb, df, sheet, 1, as_table=True, table_name='Pronóstico')

            # Change column sizes.
            change_col_sizes(sheet, col_order, col_sizes)

            if process == 'Metrics':
                mean_demand = df['Demanda'].mean()

                bias = df['Error'].mean()
                bias_perc = bias / mean_demand

                df['Error_Abs'] = df['Error'].abs()
                mae = df['Error_Abs'].mean()
                mae_perc = mae / mean_demand

                metrics_sheet = wb.create_sheet('Metricas')

                metrics_sheet.merge_cells('A1:B1')
                metrics_sheet['A1'] = 'Totales'
                metrics_sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
                metrics_sheet['A1'].font = openpyxl.styles.Font(bold=True)

                metrics_sheet['A2'] = 'Demanda Total'
                metrics_sheet['A3'] = 'Pronóstico Total'
                metrics_sheet['B2'] = df['Demanda'].sum()
                metrics_sheet['B2'].number_format = '0.00'
                metrics_sheet['B3'] = df['Pronóstico'].sum()
                metrics_sheet['B3'].number_format = '0.00'

                metrics_sheet.merge_cells('A5:B5')
                metrics_sheet['A5'] = 'Métricas'
                metrics_sheet['A5'].alignment = openpyxl.styles.Alignment(horizontal='center')
                metrics_sheet['A5'].font = openpyxl.styles.Font(bold=True)

                metrics_sheet['A6'] = 'Error Absoluto Medio'
                metrics_sheet['B6'] = mae
                metrics_sheet['B6'].number_format = '0.00'

                metrics_sheet['A7'] = 'Error Absoluto Medio (%)'
                metrics_sheet['B7'] = mae_perc
                metrics_sheet['B7'].number_format = '0.00%'

                metrics_sheet['A8'] = 'Sesgo'
                metrics_sheet['B8'] = bias
                metrics_sheet['B8'].number_format = '0.00'

                metrics_sheet['A9'] = 'Sesgo (%)'
                metrics_sheet['B9'] = bias_perc
                metrics_sheet['B9'].number_format = '0.00%'

                self.df_error_export = self.df_error_export.round(2)
                df_to_excel(wb, self.df_error_export, metrics_sheet, 11, as_table=True, table_name='Metrics')

                cols_ = ['Codigo', 'Nombre', 'Demanda', 'Pronostico', 'MAE', 'Sesgo', 'MAE (%)', 'Sesgo (%)']
                colsizes = [25, 40, 12, 13, 12, 12, 12, 12]
                change_col_sizes(metrics_sheet, cols_, colsizes)

            wb.save(os.path.join(path, file_name))
            wb.close()

        elif extension == '.csv':
            df.to_csv(os.path.join(path, file_name),
                      index=False)


if __name__ == '__main__':
    root_path = os.path.join(os.path.expanduser("~"), r'AppData\Roaming\Modulo_Demanda')

    app = Application(root_path)
