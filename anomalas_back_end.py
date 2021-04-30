import copy

import pandas as pd
import matplotlib.pyplot as plt
import os
import shelve
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

pd.options.mode.chained_assignment = None


def get_excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """

    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    result = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result) + str(row)


def create_excel_table_from_df(df: pd.DataFrame, sheet_: Worksheet, row_ini: 1, table_name):
    """Crea tabla de Excel en la hoja indicada a partir de un pandas DataFrame.

    Parametros:
    df: pandas DataFrame
    row_ini: fila inicial, por default 1
    sheet_: Worksheet object openpyxl
    table_name: nombre de la tabla"""

    col_last = get_excel_style(1, df.shape[1])[:-1]

    # Crear tabla de Excel
    tabla_excel = Table(displayName=table_name,
                        ref=f"A{row_ini}:{col_last}{df.shape[0] + row_ini}")  # nombre y tamaño

    # declarar estilos a la tabla
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)

    # asignar el estilo
    tabla_excel.tableStyleInfo = style

    # agregar tabla a la hoja
    sheet_.add_table(tabla_excel)


def df_to_excel(wb: Workbook, df: pd.DataFrame, sheet_: Worksheet, row_ini: 1, as_table: False, **kwargs):
    """Agregar pandas DataFrame a hoja de Excel.

    Parametros:
    df: pandas DataFrame
    sheet_: Worksheet object openpyxl
    row_ini: fila inicial, por default 1
    as_table: boolean, crear Tabla de Excel"""

    # Agregar dataframe de Python a Excel
    rows = dataframe_to_rows(df, index=False, header=True)

    # agregar filas a Excel
    for r_idx, row in enumerate(rows, row_ini):
        for c_idx, value in enumerate(row, 1):
            sheet_.cell(row=r_idx, column=c_idx, value=value)

    if as_table:
        try:
            table_name = kwargs['table_name']
            create_excel_table_from_df(df, sheet_, row_ini, table_name)
        except KeyError:
            raise ValueError('A table name must be specified if as_table is True.')
    try:
        for sheet in ['Sheet', 'Hoja', 'Hoja1']:
            wb.remove(wb[sheet])

    except KeyError:
        pass


class FilePathShelf:
    @staticmethod
    def close_shelf(shelf: shelve):

        shelf.close()

    def __init__(self, _path):

        # path to save the shelve files
        self._path = _path

        # open shelve
        paths_shelf = shelve.open(self._path)

        # set keys list
        self._shelve_keys = ['Export',
                             'Orders',
                             'Anomaly_Model',
                             'Temp',
                             'Export_FileName']

        # try to get value from key, if empty initialize
        for _path in self._shelve_keys:
            try:
                paths_shelf[_path]

            except KeyError:
                paths_shelf[_path] = ''

        # close shelf
        paths_shelf.close()

    def open_shelf(self):
        paths_shelf = shelve.open(self._path)

        return paths_shelf

    def write_to_shelf(self, file_name, path_):
        """Set value (path_) to key (file_name)."""

        # open saved values
        paths_shelf = self.open_shelf()

        if file_name not in self._shelve_keys:
            raise ValueError(f'You tried to save {file_name} to the dictionary. '
                             f'The accepted values are {self._shelve_keys}.')

        # set value to key
        paths_shelf[file_name] = path_

        # save and close shelf
        self.close_shelf(paths_shelf)

    def print_shelf(self):
        """Print the shelf."""

        shelf = self.open_shelf()

        for key, value in shelf.items():
            print(key, ': ', value)

            if key is None or value is None:
                pass

        # save and close shelf
        self.close_shelf(shelf)

    def send_path(self, file_name):
        """Return path from key (file_name)."""

        paths_shelf = self.open_shelf()

        if file_name not in self._shelve_keys:
            raise ValueError(f'{file_name} is not a valid file name.')

        path = paths_shelf[file_name]

        # save and close shelf
        self.close_shelf(paths_shelf)

        return path


class AnomalyApp:

    @staticmethod
    def calculate_min_max(df):

        try:
            df['Fecha'] = pd.to_datetime(df['Fecha'])
        except ValueError:
            df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y')

        group_cols = ['Cliente_Cod',
                      'Producto_Cod']

        df_25 = df.groupby(group_cols)['Cantidad'].quantile(0.05).reset_index()
        df_25 = df_25.rename(columns={'Cantidad': 'Min'})
        df_75 = df.groupby(group_cols)['Cantidad'].quantile(0.95).reset_index()
        df_75 = df_75.rename(columns={'Cantidad': 'Max'})

        df = df.merge(df_25, on=group_cols, how='left')
        df = df.merge(df_75, on=group_cols, how='left')

        df = df.drop_duplicates(subset=['Cliente_Cod', 'Producto_Cod'])
        df = df[['Cliente_Cod', 'Producto_Cod', 'Min', 'Max']]
        df.loc[df['Min'] < 0, 'Min'] = 0

        return df

    def __init__(self, path_):
        # installation path
        self.path_ = path_
        self.path_config_shelf = os.path.join(path_, 'config')
        self.path_file_paths = os.path.join(path_, 'paths')

        # attributes to save the tables created after the anomaly check
        self.df_normal = None
        self.df_anomalies = None
        self.df_missing = None

        # read the original model and save as a dataframe
        self.df_original_model = pd.read_csv('Anomalas_Modelo.csv')

        # Anomaly count
        self.anomaly_count = 0

        # initial routine
        if not self.check_if_installed():
            self.setup()

        # shelves for storing data in computer memory
        self.file_paths_shelf = FilePathShelf(self.path_file_paths)

    def setup(self):
        """
        Sets up the fixed path of the program.
        """
        if not os.path.exists(self.path_):
            print('Instalando el programa.')
            os.makedirs(self.path_)

    def check_if_installed(self):
        """
        Checks if the fixed path of the program exists or not.
        """
        if os.path.exists(self.path_):
            return True
        else:
            return False

    def set_path(self, filename, path):
        """Set path to the paths shelf."""

        self.file_paths_shelf.write_to_shelf(filename, path)

    def get_path(self, filename):
        """Get path from the paths shelf."""

        return self.file_paths_shelf.send_path(filename)

    def read_new_data(self, file):
        """
        Returns pandas dataframe with raw orders data.
        """

        # Get the Orders path from paths shelf
        path = self.file_paths_shelf.send_path(file)

        # Read the file into a pandas dataframe.
        # Change the function used depending on the extension file.
        if path.endswith('xlsx'):
            df = pd.read_excel(path)
        else:
            df = pd.read_csv(path)

        return df

    def clean_new_data(self, file):
        """Clean the raw orders data."""

        # Read the new data.
        df = self.read_new_data(file)

        # Keep columns by index.
        df = df.iloc[:, [2, 4, 5, 6, 7, 8, 9]]

        # Change column names.
        df.columns = ['Orden',
                      'Fecha',
                      'Cliente_Cod',
                      'Cliente_Nombre',
                      'Producto_Cod',
                      'Producto_Nombre',
                      'Cantidad']

        # Keep columns with non null "Cantidad" field.
        df = df[df['Cantidad'].notnull()]

        return df

    def anomaly_check(self):
        """
        Compare each client-product combination from the new orders with the anomaly model.
        Client-product combinations that are out of the range (specified by the model) must be marked as anomalies.
        Return three tables:
        df_normal = with the normal orders
        df_anomalies = with combinations that are out of range
        df_missing = with combinations that don't exist in the existing model
        """

        # Read and clean the new orders data.
        df_sales = self.clean_new_data('Orders')

        # Create a dataframe with the new orders and the minimum and maximum acceptable values for each client-product
        # combination.
        df_verification = df_sales.merge(self.df_original_model,
                                         on=['Cliente_Cod', 'Producto_Cod'],
                                         how='left')

        # If the "Cantidad" field is less than the minimum or greater than the maximum, the alert must be raised.
        df_verification['Alerta'] = 0
        df_verification.loc[(df_verification['Cantidad'] < df_verification['Min']) |
                            (df_verification['Cantidad'] > df_verification['Max']), 'Alerta'] = 1

        # Create a table with all the found alerts.
        df_anomalies = df_verification[df_verification['Alerta'] == 1]
        cols_to_drop = ['Min', 'Max', 'Alerta']
        df_anomalies.drop(columns=cols_to_drop, inplace=True)

        # The amount of rows in the alerts table is the amount of anomalies found.
        self.anomaly_count = df_anomalies.shape[0]

        # Create a table with all the missing alerts, with all the client-product combinations that don't exist in the
        # model.
        df_missing = df_verification[df_verification['Min'].isna()]
        df_missing.drop(columns=cols_to_drop, inplace=True)

        # Create a table with all the normal orders.
        df_normal = df_verification[(df_verification['Alerta'] == 0) & (df_verification['Min'].notna())]
        df_normal.drop(columns=cols_to_drop, inplace=True)

        # save tables as attributes
        self.df_normal = df_normal
        self.df_anomalies = df_anomalies
        self.df_missing = df_missing

        return df_normal, df_anomalies, df_missing

    def export_anomaly_check(self):

        # Declare excel workbook and three sheets for the three different tables
        wb = Workbook()

        df_normal = copy.deepcopy(self.df_normal)
        df_anomalies = copy.deepcopy(self.df_anomalies)
        df_missing = copy.deepcopy(self.df_missing)

        for df in [df_normal, df_anomalies, df_missing]:
            df.columns = ['Orden',
                          'Fecha',
                          'Codigo Cliente',
                          'Nombre Cliente',
                          'Codigo Producto',
                          'Nombre Producto',
                          'Cantidad']

        letters = 'ABCDEFG'
        sizes_list = [12, 12, 16, 40, 16, 40, 15]
        if not df_normal.empty:
            sheet_normal = wb.create_sheet('Correctas')
            df_to_excel(wb, df_normal, sheet_normal, 1, as_table=True, table_name='Normales')
            self.change_col_sizes(sheet_normal, letters, sizes_list)

        if not df_anomalies.empty:
            sheet_anomalies = wb.create_sheet('Anomalias')
            df_to_excel(wb, df_anomalies, sheet_anomalies, 1, as_table=True, table_name='Anomalias')
            self.change_col_sizes(sheet_anomalies, letters, sizes_list)

        if not df_missing.empty:
            sheet_missing = wb.create_sheet('Nuevos clientes')
            df_to_excel(wb, df_missing, sheet_missing, 1, as_table=True, table_name='Nuevos')
            self.change_col_sizes(sheet_missing, letters, sizes_list)

        path_ = self.file_paths_shelf.send_path('Export')
        file_name_ = self.file_paths_shelf.send_path('Export_FileName') + '.xlsx'
        full_path_ = os.path.join(path_, file_name_)
        wb.save(full_path_)
        wb.close()

    @staticmethod
    def change_col_sizes(sheet, letters: str, sizes_list: list):

        letters_list = [let for let in letters]

        for col, size in zip(letters_list, sizes_list):
            sheet.column_dimensions[col].width = size

    def update_model(self):
        """
        Update the model with new data.
        The old minimums are replaced by lower minimums and old maximums are replaced by higher maximums.
        """

        df_new = self.clean_new_data('Anomaly_Model')

        df_new = self.calculate_min_max(df_new)

        df_old = copy.deepcopy(self.df_original_model)

        df_new.columns = ['Cliente_Cod', 'Producto_Cod', 'Min_new', 'Max_new']

        df = df_old.merge(df_new, on=['Cliente_Cod', 'Producto_Cod'], how='outer')

        #
        df.loc[df['Min_new'] < df['Min'], 'Min'] = 'Min_new'
        df.loc[df['Max_new'] > df['Max'], 'Max'] = 'Max_new'

        df.loc[df['Min'].isna(), 'Min'] = 'Min_new'
        df.loc[df['Max'].isna(), 'Max'] = 'Max_new'

        # df.to_csv(os.path.join(self.path_, 'Anomalas_Modelo.csv'))

        return df
